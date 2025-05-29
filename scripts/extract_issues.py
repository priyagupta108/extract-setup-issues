import os
import requests
import datetime
import openpyxl
import time
from openpyxl.styles import Font

# -----------------------------------------------------------------------------
# Script Description:
# This script fetches GitHub issues for the repository 'actions/runner-images'
# using the REST API endpoint:
#   GET /repos/actions/runner-images/issues?state={open|closed}&since={START_DATE}&per_page=100&page={n}
# It collects issues created or updated in the last 7 months, flags special labels,
# and exports the results to an Excel file.
# -----------------------------------------------------------------------------

# Auth and repo info
TOKEN = os.getenv("GH_TOKEN")
if not TOKEN:
    raise EnvironmentError("Missing GitHub token. Please set 'GH_TOKEN' in your environment or GitHub Actions secrets.")

OWNER = "actions"
REPO = "runner-images"

# Calculate date 7 months ago
TODAY_DATE = datetime.datetime.utcnow()
START_DATE = (TODAY_DATE - datetime.timedelta(days=30 * 7)).isoformat() + "Z"
TODAY_DATE = TODAY_DATE.isoformat() + "Z"
PER_PAGE = 100

# Special labels
SPECIAL_LABELS = {
    "OS: macOS": "G",
    "OS: Ubuntu": "H",
    "OS: Windows": "I",
    "bug report": "J", 
    "feature request": "K",
    "announcement": "L"
}

headers = {
    "Authorization": f"Bearer {TOKEN}",
    "Accept": "application/vnd.github.v3+json"
}

def get_issues(state):
    issues = []
    page = 1
    while True:
        url = f"https://api.github.com/repos/{OWNER}/{REPO}/issues"
        params = {
            "state": state,
            "since": START_DATE,
            "per_page": PER_PAGE,
            "page": page
        }
        response = requests.get(url, headers=headers, params=params, timeout=90)
        if response.status_code == 401:
            raise PermissionError("❌ Unauthorized. Check if your GH_TOKEN is valid and has correct permissions.")
        response.raise_for_status()
        data = response.json()
        if not data:
            break
        for issue in data:
            created_at = issue.get("created_at")
            if created_at and START_DATE <= created_at <= TODAY_DATE and "pull_request" not in issue:
                issues.append(issue)
        page += 1
    return issues

def issues_to_excel_flagged(issues, filename="label_flags.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Label Flags"

    headers = [
        "Number", "Title", "State", "Created At", "Created Month",
        "Closed At", "Closed Month", "Days Taken", "Labels"
    ]
    label_columns = list(SPECIAL_LABELS.keys())
    headers.extend(label_columns)
    ws.append(headers)

    for issue in issues:
        labels = {lbl["name"].lower() for lbl in issue.get("labels", [])}
        created_at = issue.get("created_at", "")[:10]
        closed_at = issue.get("closed_at", "")[:10] if issue.get("closed_at") else ""

        created_date = datetime.datetime.strptime(created_at, "%Y-%m-%d") if created_at else None
        closed_date = datetime.datetime.strptime(closed_at, "%Y-%m-%d") if closed_at else None

        created_month = created_date.strftime("%b-%Y") if created_date else ""
        closed_month = closed_date.strftime("%b-%Y") if closed_date else ""
        days_taken = (closed_date - created_date).days if created_date and closed_date else ""

        issue_number = issue["number"]
        issue_url = f"https://github.com/{OWNER}/{REPO}/issues/{issue_number}"

        row = [
            issue_number,
            issue["title"],
            issue["state"],
            created_at,
            created_month,
            closed_at,
            closed_month,
            days_taken,
            ", ".join(labels)
        ]

        for label in label_columns:
            row.append("✅" if label.lower() in labels else "")

        ws.append(row)

        cell = ws.cell(row=ws.max_row, column=1)
        cell.value = issue_number
        cell.font = Font(color="0000EE", underline="single")
        cell.hyperlink = issue_url

    wb.save(filename)

if __name__ == "__main__":
    start_time = time.time()

    open_issues = get_issues("open")
    closed_issues = get_issues("closed")
    all_issues = open_issues + closed_issues

    issues_to_excel_flagged(all_issues, filename="label_flags.xlsx")

    end_time = time.time()
    elapsed_seconds = end_time - start_time
    print(f"\n✅ Script completed in {elapsed_seconds:.2f} seconds.")